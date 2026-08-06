"""Build the canonical SAE checkpoint-metrics table, figures, and workbook tab.

This script joins the reconstruction/dead-latent evaluation with the existing
loss-recovery evaluation for every available SFT and PPO SAE checkpoint.  It
creates a machine-readable CSV, two checkpoint-trajectory figures, and refreshes
the ``sae_michael`` tab in the shared collation workbook without modifying the
other collaborators' tabs.

Run from the repository root after ``06e_mse_dead_latents.py``:

    python scripts/10_build_checkpoint_metrics.py
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
RESULTS = ROOT / "results"
FIGURES = ROOT / "figures"
WORKBOOK = ROOT.parent / "collation" / "SAE_Collation.xlsx"
HIDDEN_SIZE = 896
LAYERS = [6, 12, 18, 23]

EVAL_SOURCES = [
    ("flexible", "PPO", RESULTS / "sae_eval.csv"),
    ("strict", "PPO", RESULTS / "sae_eval_strict.csv"),
    ("strict_l23_k256", "PPO", RESULTS / "sae_eval_strict_l23_k256.csv"),
    ("kl0p025", "PPO", RESULTS / "sae_eval_kl0p025.csv"),
    ("shuffled", "PPO", RESULTS / "sae_eval_shuffled.csv"),
    ("sft", "SFT", RESULTS / "sae_eval_sft.csv"),
]

PPO_COLORS = {
    "flexible": "#7B5EA7",
    "strict": "#0A014F",
    "kl0p025": "#2A9D8F",
    "shuffled": "#D66D8A",
    "strict_l23_k256": "#4E4E4E",
}


def checkpoint_step(stage: str) -> int:
    if stage == "instruct_base":
        return 0
    match = re.fullmatch(r"(?:ppo|sft)_step(\d+)", stage)
    if not match:
        raise ValueError(f"Unrecognised checkpoint label: {stage}")
    return int(match.group(1))


def read_eval_rows() -> pd.DataFrame:
    frames = []
    for chain, regime, path in EVAL_SOURCES:
        if not path.exists():
            raise FileNotFoundError(f"Missing evaluation CSV: {path}")
        frame = pd.read_csv(path)
        frame.insert(0, "chain", chain)
        frame.insert(0, "training_regime", regime)
        frames.append(frame)
    evaluations = pd.concat(frames, ignore_index=True)
    evaluations["step"] = evaluations["stage"].map(checkpoint_step)
    return evaluations


def build_metrics() -> pd.DataFrame:
    raw_paths = [RESULTS / "sae_mse_dead.csv", RESULTS / "sae_mse_dead_sft.csv"]
    missing_raw = [path for path in raw_paths if not path.exists()]
    if missing_raw:
        raise FileNotFoundError(f"Missing reconstruction metrics: {missing_raw}")
    raw = pd.concat([pd.read_csv(path) for path in raw_paths], ignore_index=True)
    evaluations = read_eval_rows()
    join_columns = ["chain", "stage", "layer", "k", "d_sae"]
    metrics = evaluations.merge(raw, how="left", on=join_columns,
                                suffixes=("", "_raw"), validate="one_to_one")
    required = ["mse", "var", "nmse_raw", "n_dead", "dead_frac"]
    missing = metrics[required].isna().any(axis=1)
    if missing.any():
        bad = metrics.loc[missing, join_columns].to_dict("records")
        raise RuntimeError(f"Missing raw metrics for {len(bad)} SAE rows: {bad}")

    metrics = metrics.rename(columns={"nmse_raw": "nmse_recomputed"})
    metrics["dead_latents_pct"] = metrics["dead_frac"] * 100.0
    metrics["frac_rec_pct"] = metrics["frac_rec"] * 100.0
    metrics["delta_loss"] = metrics["L_sae"] - metrics["L_base"]
    metrics["expansion"] = metrics["d_sae"] / HIDDEN_SIZE
    metrics["model"] = "Qwen2.5-0.5B-Instruct"
    metrics["architecture"] = "TopK"
    metrics["optimizer"] = "Adam"
    metrics["lr"] = 1e-4
    metrics["checkpoint"] = metrics["stage"]
    metrics = metrics.sort_values(
        ["training_regime", "chain", "layer", "step"],
        key=lambda col: col.map({"SFT": 0, "PPO": 1}) if col.name == "training_regime" else col,
    ).reset_index(drop=True)
    return metrics


def style_axis(axis: plt.Axes) -> None:
    axis.spines["top"].set_visible(False)
    axis.spines["right"].set_visible(False)
    axis.grid(axis="y", alpha=0.25, linewidth=0.6)


def plot_trajectories(metrics: pd.DataFrame, value: str, y_label: str,
                      title: str, filename: str) -> None:
    FIGURES.mkdir(exist_ok=True)
    fig, axes = plt.subplots(len(LAYERS), 2, figsize=(12, 12), sharex="col")
    for row, layer in enumerate(LAYERS):
        sft_ax, ppo_ax = axes[row]
        sft = metrics[(metrics.training_regime == "SFT") & (metrics.layer == layer)]
        ppo = metrics[(metrics.training_regime == "PPO") & (metrics.layer == layer)]

        for chain, group in sft.groupby("chain", sort=False):
            group = group.sort_values("step")
            sft_ax.plot(group.step, group[value], marker="o", linewidth=2,
                        color="#1F2937", label="SFT")
        for chain, group in ppo.groupby("chain", sort=False):
            group = group.sort_values("step")
            style = "--" if chain == "strict_l23_k256" else "-"
            label = "strict (K=256, L23)" if chain == "strict_l23_k256" else chain
            ppo_ax.plot(group.step, group[value], marker="o", linewidth=1.8,
                        linestyle=style, color=PPO_COLORS[chain], label=label)

        for axis, heading in ((sft_ax, "SFT"), (ppo_ax, "PPO")):
            style_axis(axis)
            axis.set_title(f"L{layer} — {heading}")
            axis.set_ylabel(y_label)
            axis.legend(frameon=False, fontsize=8, loc="best", ncol=2)
        if row == len(LAYERS) - 1:
            sft_ax.set_xlabel("SFT optimizer step")
            ppo_ax.set_xlabel("PPO optimizer step")

    fig.suptitle(title, y=0.995, fontsize=14)
    fig.tight_layout()
    fig.savefig(FIGURES / filename, dpi=300)
    plt.close(fig)


def refresh_workbook(metrics: pd.DataFrame) -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Missing collation workbook: {WORKBOOK}")
    workbook = openpyxl.load_workbook(WORKBOOK)
    if "sae_michael" in workbook.sheetnames:
        sheet = workbook["sae_michael"]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet("sae_michael", 0)

    headers = [
        "Architecture", "Expansion", "K", "Optimizer", "LR", "Checkpoint",
        "Layer", "Avg L0", "Recon Loss (MSE)", "Recon Loss (NMSE)",
        "Dead Latents %", "Delta Loss", "Frac Rec %", "Model", "Chain",
        "Training Regime",
    ]
    sheet.append(headers)
    for _, row in metrics.iterrows():
        sheet.append([
            row.architecture, int(round(row.expansion)), int(row.k), row.optimizer,
            row.lr, row.checkpoint, int(row.layer), row.mean_l0, row.mse,
            row.nmse, row.dead_latents_pct, row.delta_loss, row.frac_rec_pct,
            row.model, row.chain, row.training_regime,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [16, 11, 8, 12, 11, 18, 8, 12, 20, 21, 17, 14, 13, 27, 18, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=13):
        for cell in row:
            cell.number_format = "0.0000"
    for row in sheet.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = "0.000"
    for row in sheet.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            cell.number_format = "0.00"
    workbook.move_sheet(sheet, offset=-workbook.index(sheet))
    workbook.save(WORKBOOK)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-workbook", action="store_true")
    args = parser.parse_args()

    metrics = build_metrics()
    output = RESULTS / "sae_checkpoint_metrics.csv"
    metrics.to_csv(output, index=False, float_format="%.8f")
    plot_trajectories(
        metrics, "mse", "Raw reconstruction MSE",
        "SAE reconstruction loss over training checkpoints", "checkpoint_reconstruction_mse.png",
    )
    plot_trajectories(
        metrics, "dead_latents_pct", "Dead latents (%)",
        "SAE feature availability over training checkpoints", "checkpoint_dead_latents.png",
    )
    if not args.skip_workbook:
        refresh_workbook(metrics)
    print(f"Wrote {len(metrics)} rows -> {output}")
    print(f"Wrote figures -> {FIGURES / 'checkpoint_reconstruction_mse.png'}")
    print(f"Wrote figures -> {FIGURES / 'checkpoint_dead_latents.png'}")
    if not args.skip_workbook:
        print(f"Updated workbook -> {WORKBOOK}")


if __name__ == "__main__":
    main()
